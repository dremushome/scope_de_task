import io
import datetime
import openpyxl
from openpyxl.utils import range_boundaries

class ExcelParser:
    """
    Generic, schema-driven Excel sheet parser. Reads coordinates and extracts
    simple values or lists based strictly on the configured SCHEMA definitions.
    """

    def __init__(self, schema: dict):
        self.schema = schema
        if "sheet_name" not in schema:
            raise KeyError("Schema configuration is missing required key: 'sheet_name'")
        self.sheet_name = schema["sheet_name"]
        self.schema_type = schema.get("schema_type", "unknown")
        self.schema_version = schema.get("schema_version", "v1")
        self.sections = schema.get("sections", {})
        self.schema_sha = self._calculate_schema_sha(schema)
        self.parser_sha = self._calculate_parser_sha()

    def _calculate_schema_sha(self, schema: dict) -> str:
        import json
        import hashlib
        schema_str = json.dumps(schema, sort_keys=True)
        return hashlib.sha256(schema_str.encode('utf-8')).hexdigest()

    def _calculate_parser_sha(self) -> str:
        import inspect
        import hashlib
        parser_path = inspect.getfile(self.__class__)
        with open(parser_path, "rb") as f:
            content = f.read()
        return hashlib.sha256(content).hexdigest()

    def _resolve_boundaries(self, range_str: str, sheet) -> tuple[int, int, int, int]:
        min_c, min_r, max_c, max_r = range_boundaries(range_str)
        return (
            min_c if min_c is not None else 1,
            min_r if min_r is not None else 1,
            max_c if max_c is not None else sheet.max_column,
            max_r if max_r is not None else sheet.max_row
        )

    def clean_value(self, val):
        if val is None:
            return None
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.isoformat()[:10]
        if isinstance(val, (int, float, bool)):
            return val
        return str(val).strip()

    def parse(self, binary_content: bytes, filename: str) -> dict:
        wb = openpyxl.load_workbook(io.BytesIO(binary_content), data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Invalid spreadsheet: '{self.sheet_name}' sheet not found.")

        sheet = wb[self.sheet_name]

        parsed_data = {}
        for section in self.sections.keys():
            parsed_data[section] = {}

        # Dynamically determine sheet boundaries from SCHEMA bounds
        max_row_expected = 0
        max_col_expected = 0
        for cfg in self.sections.values():
            min_c, min_r, max_c, max_r = self._resolve_boundaries(cfg["range"], sheet)
            if max_r > max_row_expected:
                max_row_expected = max_r
            if max_c > max_col_expected:
                max_col_expected = max_c

        # Find metrics header row index for dynamic coordinate validation
        metrics_header_r_idx = None
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if len(row) > 1 and row[1] == '[Scope Credit Metrics]':
                metrics_header_r_idx = r_idx
                break

        # Scan for unexpected cells to enforce data contract
        unexpected_cells = {}
        for r_idx in range(1, sheet.max_row + 1):
            for c_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                val = cell.value
                if val is None:
                    continue
                
                # Check bounds
                if r_idx > max_row_expected or c_idx > max_col_expected:
                    unexpected_cells[cell.coordinate] = val
                    continue
                
                is_expected = False
                if c_idx == 2: # Column B contains labels
                    is_expected = True
                else:
                    for section, section_cfg in self.sections.items():
                        min_c, min_r, max_c, max_r = self._resolve_boundaries(section_cfg["range"], sheet)
                        if min_r <= r_idx <= max_r:
                            if min_c <= c_idx <= max_c:
                                # Check if it is a configured ignored cell value
                                ignored_cells = section_cfg.get("ignored_cell_values", {})
                                val_str = str(val).strip()
                                was_ignored = False
                                for val_key, ignored_range in ignored_cells.items():
                                    if val_str == val_key:
                                        min_ic, min_ir, max_ic, max_ir = self._resolve_boundaries(ignored_range, sheet)
                                        if min_ir <= r_idx <= max_ir and min_ic <= c_idx <= max_ic:
                                            is_expected = True
                                            was_ignored = True
                                            break
                                if was_ignored:
                                    break
                                
                                # Otherwise check if row label matches expected categories
                                row_label_cell = sheet.cell(row=r_idx, column=2)
                                row_label = str(row_label_cell.value).strip() if row_label_cell.value else None
                                if row_label in section_cfg.get("key_value", []) or row_label in section_cfg.get("named_list", []):
                                    is_expected = True
                                    break
                
                if not is_expected:
                    unexpected_cells[cell.coordinate] = val

        parsed_data['unexpected_cells'] = unexpected_cells

        # Read and parse rows based purely on simple/list definitions
        rows = list(sheet.iter_rows(values_only=True))
        for r_idx, row in enumerate(rows, start=1):
            if len(row) < 2 or not isinstance(row[1], str):
                continue

            label = row[1].strip()

            # Find matching section in SCHEMA
            for section, section_cfg in self.sections.items():
                min_c, min_r, max_c, max_r = self._resolve_boundaries(section_cfg["range"], sheet)
                if min_r <= r_idx <= max_r:
                    # Parse simple fields
                    if label in section_cfg.get("key_value", []):
                        parsed_data[section][label] = self.clean_value(row[min_c - 1])
                    # Parse list fields
                    elif label in section_cfg.get("named_list", []):
                        row_slice = []
                        for col_idx in range(min_c, max_c + 1):
                            is_ignored = False
                            ignored_cells = section_cfg.get("ignored_cell_values", {})
                            for val_str, ignored_range in ignored_cells.items():
                                min_ic, min_ir, max_ic, max_ir = self._resolve_boundaries(ignored_range, sheet)
                                if min_ir <= r_idx <= max_ir and min_ic <= col_idx <= max_ic:
                                    is_ignored = True
                                    break
                            if not is_ignored:
                                row_slice.append(row[col_idx - 1])
                        
                        cleaned_slice = [self.clean_value(x) for x in row_slice]
                        # Trim trailing None values
                        while cleaned_slice and cleaned_slice[-1] is None:
                            cleaned_slice.pop()
                        parsed_data[section][label] = cleaned_slice

        # Embed metadata and run validation
        validation_errors = self.validate(parsed_data)
        passed = (len(validation_errors) == 0)

        parsed_data["_meta"] = {
            "filename": filename,
            "schema_type": self.schema_type,
            "schema_version": self.schema_version,
            "schema_sha": self.schema_sha,
            "parser_sha": self.parser_sha,
            "data_quality_passed": passed,
            "data_quality_errors": validation_errors
        }

        return parsed_data

    def validate(self, parsed_data: dict) -> list:
        errors = []
        
        # 1. Check for unexpected cells found during parsing
        unexpected_cells = parsed_data.get('unexpected_cells', {})
        for cell_ref, val in unexpected_cells.items():
            errors.append(f"Unexpected data point found in cell {cell_ref}: '{val}'")

        # 2. Check for missing required key_value fields in each section
        for section, section_cfg in self.sections.items():
            optional_keys = section_cfg.get("optional_keys", [])
            section_data = parsed_data.get(section, {})
            
            for key in section_cfg.get("key_value", []):
                val = section_data.get(key)
                if (val is None or val == "") and key not in optional_keys:
                    errors.append(f"Missing required field: '{key}' in section '{section}'")
                    
        return errors
